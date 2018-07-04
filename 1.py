import pymysql
import openpyxl
db = pymysql.connect("localhost","root","123456","ex1")
cursor = db.cursor()
cursor.execute("SHOW databases")
data = cursor.fetchall()
print(data)
cursor.execute("USE ex1")
data = cursor.fetchall()
print(data)
cursor.execute("SHOW tables")
data = cursor.fetchall()
print(data)
cursor.execute("select * from student")
data = cursor.fetchall()
print(data)
print(type(data))
print(len(data))
print(len(data[0]))
# db.close()
# wb = openpyxl.load_workbook("1.xlsx")
# print(wb.sheetnames)
# sheet = wb["Sheet1"]
# for i in sheet["C"]:
#     print(i.value,end = " ")
# temp2 = openpyxl.Workbook()
# sheet = wb.active
# sheet.title = "new sheet"
# sheet['C3'] = 'Hello world'
# for i in range(10):
#     sheet["A%d" % (i+1)].value = i+1
# wb.save("2.xlsx")
wb = openpyxl.Workbook()
temp2 = wb.active
temp2.title = "2"
h = len(data)
w = len(data[0])
for i in range(h):
    for j in range(w):
        temp2.cell(row = i+1,column = j+1).value = data[i][j]
wb.save("3.xlsx")



